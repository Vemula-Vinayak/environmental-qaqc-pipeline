"""
build_excel_deliverable.py
Builds the client-deliverable-style Excel workbook:
  - Validation Summary
  - Regulatory Exceedance Summary
  - Holding Time Violations detail
  - Full Validated Dataset
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

OUT = "/home/claude/envqaqc/outputs"

df = pd.read_csv(f"{OUT}/validated_dataset.csv")
validation_summary = pd.read_csv(f"{OUT}/validation_summary.csv")
exceedance_summary = pd.read_csv(f"{OUT}/exceedance_summary.csv")
holding_violations = df[df["flag_holding_time_violation"]][
    ["sample_id", "site_name", "location_id", "analyte", "collection_date",
     "analysis_date", "holding_time_days", "analyte_category"]
].sort_values("holding_time_days", ascending=False)

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F2937")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")


def write_df(ws, df_in, start_row=3, title=None):
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for j, col in enumerate(df_in.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for i, row in enumerate(df_in.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
    for j, col in enumerate(df_in.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df_in[col].astype(str)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 12), 40)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


wb = Workbook()

# --- Sheet 1: Validation Summary ---
ws1 = wb.active
ws1.title = "Validation Summary"
write_df(ws1, validation_summary, title="Environmental Data QA/QC — Validation Summary")
# conditional shading on fail_rate_pct
for r in range(4, 4 + len(validation_summary)):
    fail_rate = ws1.cell(row=r, column=6).value
    fill = FAIL_FILL if (fail_rate is not None and fail_rate > 10) else PASS_FILL
    for c in range(1, 7):
        ws1.cell(row=r, column=c).fill = fill

overall_pass = (df["overall_qaqc_status"] == "PASS").mean() * 100
ws1.cell(row=4 + len(validation_summary) + 2, column=1,
         value=f"Overall record-level QA/QC pass rate: {overall_pass:.2f}%  (n = {len(df)} records)").font = Font(bold=True, name="Arial")

# --- Sheet 2: Regulatory Exceedance Summary ---
ws2 = wb.create_sheet("Regulatory Exceedances")
write_df(ws2, exceedance_summary, title="Regulatory Exceedance Summary — EPA MCL Comparison by Site & Analyte")

# --- Sheet 3: Holding Time Violations ---
ws3 = wb.create_sheet("Holding Time Violations")
write_df(ws3, holding_violations, title="Holding Time Violations — Collection-to-Analysis Lag vs. EPA Standard")

# --- Sheet 4: Full Validated Dataset ---
ws4 = wb.create_sheet("Full Validated Dataset")
display_cols = [c for c in df.columns]
write_df(ws4, df[display_cols], title="Full Validated Dataset (All Flags Applied)")

wb.save(f"{OUT}/Environmental_QAQC_Deliverable.xlsx")
print("Excel deliverable saved.")
